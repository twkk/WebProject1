
from bs4 import BeautifulSoup
import requests
import pymysql
import openpyxl
from openpyxl.styles import Font
import gspread
from oauth2client.service_account import ServiceAccountCredentials


class Stock:
    def __init__(self, *stock_numbers):
        self.stock_numbers = stock_numbers

    def scrape(self):

        result = list()

        for stock_number in self.stock_numbers:

            response = requests.get(
                "https://tw.stock.yahoo.com/q/q?s=" + stock_number)
            soup = BeautifulSoup(response.text.replace("加到投資組合", ""), "lxml")

            stock_date = soup.find(
                "font", {"class": "tt"}).getText().strip()[-9:]  # 資料日期

            tables = soup.find_all("table")[2]  # 取得網頁中第三個表格
            tds = tables.find_all("td")[0:11]  # 取得表格中1到10格

            result.append((stock_date,) +
                          tuple(td.getText().strip() for td in tds))
        return result

    def save(self, stocks):

        db_settings = { "host": "127.0.0.1",
            "port": 3306,"user": "root", "password": "abrazo25",
            "db": "stock","charset": "utf8"
        }

        try:
            conn = pymysql.connect(**db_settings)

            with conn.cursor() as cursor:
                sql = """INSERT INTO market(
                                market_date,
                                stock_name,
                                market_time,
                                final_price,
                                buy_price,
                                sell_price,
                                ups_and_downs,
                                lot,
                                yesterday_price,
                                opening_price,
                                highest_price,
                                lowest_price)
                         VALUES(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""

                for stock in stocks:
                    cursor.execute(sql, stock)
                conn.commit()

        except Exception as ex:
            print("Exception:", ex)

    def export(self, stocks):
        wb = openpyxl.Workbook()
        sheet = wb.create_sheet("Yahoo股市", 0)

        response = requests.get(
            "https://tw.stock.yahoo.com/q/q?s=2451")
        soup = BeautifulSoup(response.text, "lxml")

        tables = soup.find_all("table")[2]
        ths = tables.find_all("th")[0:11]
        titles = ("資料日期",) + tuple(th.getText() for th in ths)
        sheet.append(titles)

        for index, stock in enumerate(stocks):
            sheet.append(stock)

            if "△" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='FF0000')
            elif "▽" in stock[6]:
                sheet.cell(row=index+2, column=7).font = Font(color='00A600')

        wb.save("yahoostock.xlsx")

    def gsheet(self, stocks):
        scopes = ["https://spreadsheets.google.com/feeds"]
        credentials = ServiceAccountCredentials.from_json_keyfile_name(
            "C:\\Users\\keith\\source\\repos\\WebProject1\\credentials.json", scopes) 
        client = gspread.authorize(credentials)
        sheet = client.open_by_key(
            "19g37kkpLm_x1gkQ00AAYGo0NAfGQ5tK3VNleyEAfyU4").sheet1

        response = requests.get(
            "https://tw.stock.yahoo.com/q/q?s="+ str(stocks))
        soup = BeautifulSoup(response.text, "lxml")

        tables = soup.find_all("table")[2]
        ths = tables.find_all("th")[0:11]
        titles = ("資料日期",) + tuple(th.getText() for th in ths)
        sheet.append_row(titles, 1)

        for stock in stocks:
            sheet.append_row(stock)

#os.chdir("C://Money")
stock = Stock(
'1240','1258', '1259', '1264','1268','1333','1336','1565','1569','1570','1580','1584','1586','1591','1593','1595','1597','1599','1742','1752','1777','1781','1784','1785','1788','1796','1799','1813','1815'
)  # 建立Stock物件
stock.gsheet(stock.scrape()) 

stock = Stock(
'2035','2061','2063','2064','2065','2066','2067','2070','2221','2230','2235','2596','2640','2641','2643','2718','2719','2724','2726','2729','2732','2734','2736','2740','2743','2745','2752','2916','2924','2926','2928','2937'
)
stock.gsheet(stock.scrape()) 
stock = Stock(
'3064','3066','3067','3071','3073','3078','3081','3083','3085','3086','3088','3089','3092','3093','3095','3105','3114','3115','3118','3122','3128','3131','3141','3144','3147','3152','3162','3163','3169','3171','3176','3178','3188','3191','3202','3205','3206','3207','3211','3213','3217','3218','3219','3221','3224','3226','3227','3228','3230','3232','3234','3236','3252','3259','3260','3264','3265','3268','3272','3276','3284','3285','3287','3288','3289','3290','3293','3294','3297','3303','3306','3310','3313','3317','3322','3323','3324','3325','3332','3339','3354','3360','3362','3363','3372','3373','3374','3379','3388','3390','3402','3426','3434','3438','3441','3444','3455','3465','3466','3479','3483','3484','3489','3490','3491','3492','3498','3499','3508','3511','3512','3516','3520','3521','3522','3523','3526','3527','3529','3531','3537','3540','3541','3546','3548','3551','3552','3555','3556','3558','3564','3567','3570','3577','3580','3581','3587','3594','3597','3609','3611','3615','3623','3624','3625','3628','3629','3630','3631','3632','3642','3646','3652','3663','3664','3666','3672','3675','3680','3684','3685','3687','3689','3691','3693','3707','3709','3710','3713'
)
stock.gsheet(stock.scrape()) 
stock = Stock(
 '4102','4105','4107','4109','4111','4113','4114','4116','4120','4121','4123','4126','4127','4128','4129','4130','4131','4138','4139','4147','4152','4153','4154','4157','4160','4161','4162','4163','4167','4168','4171','4173','4174','4175','4183','4188','4192','4198','4205','4207','4303','4304','4305','4401','4402','4406','4413','4416','4417','4419','4420','4429','4430','4432','4433','4502','4503','4506','4510','4513','4523','4527','4528','4529','4530','4533','4534','4535','4538','4541','4542','4543','4549','4550','4554','4556','4561','4563','4568','4580','4609','4702','4706','4707','4711','4712','4714','4716','4721','4726','4728','4729','4735','4736','4741','4743','4744','4745','4747','4754','4760','4767','4803','4804','4806','4903','4905','4907','4908','4909','4911','4924','4931','4933','4939','4944','4946','4950','4953','4966','4971','4972','4973','4974','4979','4987','4991','4995'
)
stock.gsheet(stock.scrape()) 
stock = Stock(
'5009','5011','5013','5014','5015','5016','5102','5201','5202','5205','5206','5209','5210','5211','5212','5213','5220','5223','5227','5230','5245','5251','5263','5272','5274','5276','5278','5281','5287','5289','5291','5299','5301','5302','5304','5306','5309','5310','5312','5314','5315','5321','5324','5328','5340','5344','5345','5347','5348','5349','5351','5353','5355','5356','5364','5371','5381','5383','5386','5392','5398','5403','5410','5425','5426','5432','5438','5439','5443','5450','5452','5455','5457','5460','5464','5465','5468','5474','5475','5478','5481','5483','5487','5488','5489','5490','5493','5498','5508','5511','5512','5514','5516','5520','5523','5529','5530','5536','5543','5601','5603','5604','5609','5701','5703','5704','5820','5864','5878','5902','5903','5904','5905'
)
stock.gsheet(stock.scrape()) 
stock = Stock(
'6015','6016','6020','6021','6023','6026','6101','6103','6104','6109','6111','6113','6114','6118','6121','6122','6123','6124','6125','6126','6127','6129','6130','6134','6138','6140','6143','6144','6146','6147','6148','6150','6151','6154','6156','6158','6160','6161','6163','6167','6169','6170','6171','6173','6174','6175','6179','6180','6182','6185','6186','6187','6188','6190','6194','6195','6198','6199','6203','6204','6207','6208','6210','6212','6217','6218','6219','6220','6221','6222','6223','6227','6228','6229','6231','6233','6234','6236','6237','6240','6241','6242','6244','6245','6246','6247','6248','6259','6261','6263','6264','6265','6266','6270','6274','6275','6276','6279','6284','6287','6290','6291','6292','6294','6404','6411','6417','6418','6419','6425','6426','6432','6435','6438','6441','6446','6457','6461','6462','6465','6469','6470','6472','6482','6485','6486','6488','6492','6494','6496','6499','6506','6508','6509','6510','6512','6514','6516','6523','6527','6530','6532','6535','6538','6542','6547','6548','6556','6560','6561','6568','6569','6570','6574','6576','6577','6578','6588','6589','6590','6593','6594','6596','6603','6609','6612','6613','6615','6616','6624','6629','6640','6642','6643','6649','6654','6662','6664','6667','6679','6680','6683','6690','6697','6716','6732','6803'
)
stock = Stock(
'7402','8024','8027','8032','8034','8038','8040','8042','8043','8044','8047','8048','8049','8050','8054','8059','8064','8066','8067','8068','8069','8071','8074','8076','8077','8080','8083','8084','8085','8086','8087','8088','8091','8092','8093','8096','8097','8099','8107','8109','8111','8121','8147','8155','8171','8176','8182','8183','8234','8240','8255','8277','8279','8284','8289','8291','8299','8342','8349','8354','8358','8383','8390','8401','8403','8406','8409','8410','8415','8416','8418','8420','8421','8423','8424','8426','8431','8432','8433','8435','8436','8437','8440','8444','8446','8450','8455','8472','8476','8477','8489','8905','8906','8908','8916','8917','8921','8923','8924','8927','8928','8929','8930','8931','8932','8933','8934','8935','8936','8937','8938','8941','8942','9949','9950','9951','9960','9962'
)
stock.gsheet(stock.scrape())  # 將爬取的果寫入Google Sheet工作表
# stock.export(stock.scrape())  # 將爬取的結果匯出成Excel檔案
# stock.save(stock.scrape())  # 將爬取的結果存入MySQL資料庫中



